# api/notify_views.py
import csv
import io
import logging
from typing import Dict, Any, List
import requests
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone
from rest_framework.decorators import (
    api_view,
    permission_classes,
    authentication_classes,
)
from django.http import HttpResponse
from .utils_wa import normalize_wa

from django.db import transaction
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook
from .models import WaContact, WaInbox
from .models import Curso, Enrol, UserProfile, Horario, PendingRole, LoginPIN, MZSetting
log = logging.getLogger(__name__)

from datetime import date
from calendar import monthrange
from io import BytesIO
import tempfile
from pathlib import Path

import re
import os
from .models import WaContact
from django.contrib.auth import get_user_model
from django.core.cache import cache

from django.db.models import Q

def _teacher_profile_by_from(from_wa: str):
    wa9 = normalize_wa(from_wa)
    if not wa9:
        return None
    return (UserProfile.objects
            .filter(wa=wa9, user__is_staff=True)   # staff = teacher
            .select_related("user")
            .first())

def _teacher_course_codes(tprof: UserProfile) -> list[str]:
    return list(
        Enrol.objects
        .filter(user=tprof.user, role="teacher")
        .exclude(codigo="")
        .values_list("codigo", flat=True)
        .order_by("codigo")
    )

from datetime import timedelta

def _course_span_from_horario(codigo: str):
    """
    Возвращает (start_date, end_date) как MIN/MAX(dia) по Horario (только tipo=curso).
    """
    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True)))

    start = qs.order_by("dia").values_list("dia", flat=True).first()
    end   = qs.order_by("dia").values_list("dia", flat=True).last()
    return start, end


def _horario_time_key(h: Horario) -> str:
    hi = h.hora_inicio.strftime("%H:%M") if h.hora_inicio else ""
    hf = h.hora_fin.strftime("%H:%M") if h.hora_fin else ""
    return f"{hi}–{hf}" if hi and hf else ""


def _course_time_segments(codigo: str) -> list[dict]:
    """
    Делает tramos по времени (как экспорт): склеивает подряд идущие дни
    с одинаковым временем HH:MM–HH:MM.
    Возвращает список: [{from:date, to:date, time:"09:00–14:00"}...]
    """
    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True))
          .order_by("dia", "hora_inicio"))

    # строим: dia -> most common time of day (если несколько записей в день)
    by_day = {}
    for h in qs.only("dia", "hora_inicio", "hora_fin"):
        if not h.dia:
            continue
        t = _horario_time_key(h)
        if not t:
            continue
        by_day.setdefault(h.dia, []).append(t)

    if not by_day:
        return []

    # choose most common time per day
    day_times = []
    for d in sorted(by_day.keys()):
        from collections import Counter
        t = Counter(by_day[d]).most_common(1)[0][0]
        day_times.append((d, t))

    # merge contiguous days with same time
    segs = []
    cur_from, cur_to, cur_t = day_times[0][0], day_times[0][0], day_times[0][1]
    for d, t in day_times[1:]:
        if t == cur_t and d == (cur_to + timedelta(days=1)):
            cur_to = d
        else:
            segs.append({"from": cur_from, "to": cur_to, "time": cur_t})
            cur_from, cur_to, cur_t = d, d, t
    segs.append({"from": cur_from, "to": cur_to, "time": cur_t})
    return segs


def _teacher_welcome_text(request, tprof: UserProfile, codes: list[str]) -> str:
    name = (tprof.display_name or tprof.build_display_name() or "profe").strip()
    link = request.build_absolute_uri("/acceder/?tab=profile")

    if not codes:
        return (
            f"Hola {name} 👋\n"
            "Eres *DOCENTE* en MEATZE.\n"
            f"Perfil: {link}\n\n"
            "Ahora mismo no tienes cursos asignados.\n"
            "Si esto es un error, avisa a administración."
        )

    lines = [f"Hola {name} 👋",
             "Eres *DOCENTE* en MEATZE.",
             f"Perfil: {link}",
             "",
             "Tus cursos:"]
    for c in codes:
        lines.append(f"• {c}")

    lines += [
        "",
        "Pulsa un curso (botón) y te envío calendario por meses + imágenes."
    ]

    return "\n".join(lines)

def _teacher_course_info_text(request, codigo: str) -> str:
    c = Curso.objects.filter(codigo=codigo).first()
    if not c:
        return f"❌ No encuentro el curso *{codigo}*."

    start, end = _course_span_from_horario(codigo)
    start_txt = start.isoformat() if start else "—"
    end_txt   = end.isoformat() if end else "—"

    segs = _course_time_segments(codigo)
    if segs:
        # основной "default" — самый частый time по всем сегментам
        from collections import Counter
        common = Counter([s["time"] for s in segs]).most_common(1)[0][0]
        horario_txt = common
    else:
        horario_txt = "—"

    out = [
        f"*{c.codigo}* — {c.titulo}",
        f"Fechas: {start_txt} → {end_txt}",
        f"Horario: {horario_txt}",
        "",
        "🧩 Cambios (tramos):"
    ]

    if not segs:
        out.append("— (sin horario)")
    else:
        for s in segs[:12]:
            out.append(f"• {s['from'].isoformat()} → {s['to'].isoformat()} · {s['time']}")

    out += ["", "📅 Calendario por meses (como export):"]
    return "\n".join(out)


def _course_months_by_horario(codigo: str):
    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True))
          .order_by("dia"))

    first = qs.values_list("dia", flat=True).first()
    last  = qs.values_list("dia", flat=True).last()
    if not first or not last:
        return []

    y, m = first.year, first.month
    months = []
    while (y < last.year) or (y == last.year and m <= last.month):
        months.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return months

def _month_summary_lines(codigo: str, year: int, month: int):
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])

    qs = (Horario.objects
          .filter(curso__codigo=codigo, dia__gte=first_day, dia__lte=last_day)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True))
          .order_by("dia", "hora_inicio"))

    rows = []
    for h in qs:
        hi = h.hora_inicio.strftime("%H:%M") if h.hora_inicio else ""
        hf = h.hora_fin.strftime("%H:%M") if h.hora_fin else ""
        if hi and hf:
            rows.append((h.dia.isoformat(), f"{hi}–{hf}"))

    if not rows:
        return ["— (sin clases este mes)"]

    times = [t for _, t in rows]
    from collections import Counter
    default_time, _ = Counter(times).most_common(1)[0]

    # диапазон дат месяца
    dmin = rows[0][0]
    dmax = rows[-1][0]

    # исключения (дни с другим временем)
    exceptions = [(d, t) for d, t in rows if t != default_time]
    exc_lines = []
    for d, t in exceptions[:12]:  # лимит, чтобы WA не раздувать
        exc_lines.append(f"   • {d} → {t}")

    line0 = f"{year}-{month:02d}: {dmin} → {dmax} · {default_time}"
    if exc_lines:
        return [line0, "  Cambios de horario:", *exc_lines]
    return [line0]


def short_mod(s: str, n: int = 10) -> str:
    s = (s or "").strip()
    if not s:
        return "—"
    return (s[:n] + "…") if len(s) > n else s


def build_legend_for_month(codigo: str, year: int, month: int) -> str:
    """
    Легенда: уникальные модули месяца, сокращённые до 10 символов.
    Только tipo=curso (без practica).
    """
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])

    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True))
          .filter(dia__gte=first_day, dia__lte=last_day)
          .order_by("dia", "hora_inicio"))

    mods = []
    seen = set()
    for h in qs:
        m = short_mod(h.modulo, 10)
        if m not in seen and m != "—":
            seen.add(m)
            mods.append(m)

    if not mods:
        return "Leyenda: (sin módulos este mes)"
    return "Leyenda: " + " · ".join(mods[:20])  # ограничим чтобы не раздувать

def build_month_calendar_html(codigo: str, year: int, month: int) -> str:
    c = Curso.objects.filter(codigo=codigo).first()
    title = f"{codigo} — {(c.titulo or '').strip()}" if c else codigo

    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])

    # занятия только curso (без practica)
    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True))
          .filter(dia__gte=first_day, dia__lte=last_day)
          .order_by("dia", "hora_inicio"))

    by_day = {}
    for h in qs:
        key = h.dia.isoformat()
        by_day.setdefault(key, []).append(h)

    # понедельник=0..воскресенье=6
    start_wd = (first_day.weekday())  # Mon=0
    days_in_month = last_day.day

    month_name = first_day.strftime("%B %Y")  # если хочешь ES — можно руками мапу

    # HTML
    # ВАЖНО: делаем фиксированную ширину, крупный шрифт, контраст, чтобы PNG было читаемо.
    html = f"""
<!doctype html>
<html>
<head>
<meta charset="utf-8" />
<style>
  body {{
    margin: 0;
    font-family: Inter, Arial, sans-serif;
    background: #0b1220;
    color: #e5e7eb;
  }}
  .wrap {{
    width: 1100px;
    padding: 28px 28px 22px;
  }}
  .top {{
    display:flex; justify-content:space-between; align-items:flex-end;
    margin-bottom: 16px;
  }}
  .h1 {{
    font-size: 26px; font-weight: 800; letter-spacing: .02em;
    margin:0;
  }}
  .h2 {{
    font-size: 16px; opacity: .9; margin: 4px 0 0 0;
  }}
  .badge {{
    font-size: 13px;
    border: 1px solid rgba(148,163,184,.45);
    background: rgba(15,23,42,.9);
    padding: 8px 12px;
    border-radius: 999px;
    white-space: nowrap;
  }}
  .grid {{
    display:grid;
    grid-template-columns: repeat(7, 1fr);
    gap: 10px;
  }}
  .dow {{
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: .12em;
    opacity: .75;
    padding: 6px 10px;
  }}
  .cell {{
    border-radius: 16px;
    border: 1px solid rgba(148,163,184,.22);
    background: rgba(15,23,42,.75);
    padding: 10px 10px 10px;
    min-height: 120px;
    position: relative;
    overflow: hidden;
  }}
  .day {{
    font-size: 13px;
    opacity: .9;
    font-weight: 700;
  }}
  .items {{
    margin-top: 8px;
    display:flex; flex-direction:column; gap:6px;
  }}
  .it {{
    border-radius: 12px;
    padding: 8px 9px;
    background: rgba(59,130,246,.16);
    border: 1px solid rgba(59,130,246,.22);
    font-size: 12px;
    line-height: 1.25;
  }}
  .it b {{ font-weight: 800; }}
  .muted {{ opacity:.8; }}
  .empty {{
    border: 1px dashed rgba(148,163,184,.22);
    background: rgba(2,6,23,.28);
  }}
  .legend {{
    margin-top: 14px;
    font-size: 12px;
    opacity: .85;
  }}
</style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <p class="h1">{month_name}</p>
        <p class="h2">{title}</p>
      </div>
      <div class="badge">MEATZE · Docentes</div>
    </div>

    <div class="grid" style="margin-bottom:10px;">
      <div class="dow">LUN</div><div class="dow">MAR</div><div class="dow">MIÉ</div><div class="dow">JUE</div><div class="dow">VIE</div><div class="dow">SÁB</div><div class="dow">DOM</div>
    </div>

    <div class="grid">
"""
    # пустые клетки до 1 числа
    for _ in range(start_wd):
        html += '<div class="cell empty"></div>'

    # дни месяца
    for d in range(1, days_in_month + 1):
        dt = date(year, month, d)
        key = dt.isoformat()
        items = by_day.get(key, [])

        html += f'<div class="cell"><div class="day">{d:02d}</div><div class="items">'
        if not items:
            html += '<div class="muted" style="margin-top:6px;font-size:12px;">—</div>'
        else:
            for h in items[:4]:  # чтобы не забивать клетку
                mod = short_mod(h.modulo, 10)
                aula = (h.aula or "—").strip()
                hh1 = h.hora_inicio.strftime("%H:%M")
                hh2 = h.hora_fin.strftime("%H:%M")
                html += f'<div class="it"><b>{hh1}-{hh2}</b> · {aula}<br><span class="muted">{mod}</span></div>'
        html += '</div></div>'

    # добивка до конца недели
    total_cells = start_wd + days_in_month
    tail = (7 - (total_cells % 7)) % 7
    for _ in range(tail):
        html += '<div class="cell empty"></div>'

    legend = build_legend_for_month(codigo, year, month)
    html += f"""
    </div>
    <div class="legend">{legend}</div>
  </div>
</body>
</html>
"""
    return html


def render_html_to_png_bytes(html: str, width: int = 1100, height: int = 0, device_scale: float = 2.0) -> bytes:
    """
    Рендерим HTML в PNG через Playwright.
    device_scale=2.0 даёт "retina" качество.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:
        raise RuntimeError("Playwright not installed. Install: pip install playwright && playwright install chromium") from e

    with sync_playwright() as p:
        browser = p.chromium.launch(args=["--no-sandbox"])
        page = browser.new_page(viewport={"width": width, "height": 800}, device_scale_factor=device_scale)

        page.set_content(html, wait_until="networkidle")

        # Если высоту не задаём — делаем full_page
        png = page.screenshot(type="png", full_page=True)
        browser.close()
        return png

@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def teacher_calendar_media(request):
    """
    POST /meatze/v5/notify/teacher-calendar
    body: { "codigo":"IFCT0309", "year":2026, "month":1 }
    Возвращает { ok:true, url:"https://.../media/..." , legend:"..." }
    (Пока без auth — но лучше потом защитить, если надо)
    """
    codigo = (request.data.get("codigo") or "").strip().upper()
    year = int(request.data.get("year") or 0)
    month = int(request.data.get("month") or 0)

    if not codigo or year < 2000 or month < 1 or month > 12:
        return Response({"ok": False, "message": "bad_params"}, status=400)

    html = build_month_calendar_html(codigo, year, month)
    legend = build_legend_for_month(codigo, year, month)

    try:
        png_bytes = render_html_to_png_bytes(html, width=1100, device_scale=2.0)
    except Exception as e:
        log.exception("teacher_calendar_media render failed")
        return Response({"ok": False, "message": "render_failed", "detail": str(e)}, status=500)

    filename = f"{codigo}_{year}{month:02d}_calendar.png"
    path = default_storage.save(f"wa_calendar/{filename}", ContentFile(png_bytes))
    url = request.build_absolute_uri(default_storage.url(path))

    return Response({"ok": True, "url": url, "legend": legend})


def _send_teacher_month_calendar(request, to_wa: str, codigo: str, year: int, month: int):
    html = build_month_calendar_html(codigo, year, month)
    legend = build_legend_for_month(codigo, year, month)
    try:
        png_bytes = render_html_to_png_bytes(html, width=1100, device_scale=2.0)

        filename = f"{codigo}_{year}{month:02d}_calendar.png"
        path = default_storage.save(f"wa_calendar/{filename}", ContentFile(png_bytes))
        url = request.build_absolute_uri(default_storage.url(path))

        # caption короткий, легенду отдельным сообщением (надежнее)
        wa_send_image(to_wa, url, caption=f"{codigo} · {year}-{month:02d}")
        wa_send_text(to_wa, legend)
    except Exception as e:
        log.exception("Calendar render failed: %s", e)
        wa_send_text(to_wa, "No puedo generar el calendario ahora mismo (render).")
        wa_send_text(to_wa, legend)  # хотя бы легенду
        return


from collections import Counter

def _course_time_summary(codigo: str):
    qs = (Horario.objects
          .filter(curso__codigo=codigo)
          .filter(Q(tipo="") | Q(tipo__isnull=True) | Q(tipo="curso"))
          .filter(Q(grupo="") | Q(grupo__isnull=True)))

    pairs = []
    for h in qs.only("hora_inicio", "hora_fin"):
        if h.hora_inicio and h.hora_fin:
            pairs.append((h.hora_inicio.strftime("%H:%M"), h.hora_fin.strftime("%H:%M")))

    if not pairs:
        return "—"

    # берём наиболее частое “с-по”
    (a,b), _ = Counter(pairs).most_common(1)[0]
    return f"{a}–{b}"


def sync_teacher_contact(profile: "UserProfile"):
    wa = normalize_wa(profile.wa)
    if not wa:
        return None

    # имя для WaContact
    name = (profile.display_name or profile.build_display_name() or profile.user.get_full_name() or "").strip()

    # loc можно оставить пустым или брать из Enrol/Curso (если у тебя есть правило)
    obj, _ = WaContact.objects.update_or_create(
        wa=wa,
        defaults={
            "name": name,
            "active": True,
        }
    )
    return obj
    

def find_teacher_by_wa(wa9: str):
    wa9 = normalize_wa(wa9)
    if not wa9:
        return None
    return (UserProfile.objects
            .select_related("user")
            .filter(wa=wa9, user__is_staff=True)
            .first())


def default_reply_text(profile_name: str) -> str:
    return (
        "Hola 👋 Gracias por escribir a MEATZE.\n"
        "Hemos recibido tu mensaje. Te responderemos en breve."
    )


# ========== ADMIN GUARD (аналог mz_admin_ok / mz_admin_guard) ==========

def _admin_ok(request) -> bool:
    token = (
        request.headers.get("X-MZ-Admin")
        or request.GET.get("adm")
        or request.data.get("adm")
        or ""
    )
    expected = getattr(settings, "MEATZE_ADMIN_PASS", "")
    return bool(token and expected and token == expected)


def _require_admin(request):
    if _admin_ok(request):
        return None
    return Response(
        {"message": "No autorizado (admin token requerido)."},
        status=status.HTTP_403_FORBIDDEN,
    )


# ========== WA HELPERS (эквивалент mz_wa_msisdn, mz_wa_api, шаблоны) ==========

def wa_msisdn(num: str) -> str:
    """
    Аналог mz_wa_msisdn:
    - оставляем только цифры
    - если начинается с 34 и длина > 9 — считаем ок
    - если длина == 9 → добавляем префикс 34
    """
    digits = "".join(ch for ch in str(num) if ch.isdigit())
    if digits == "":
        return ""
    if digits.startswith("34") and len(digits) > 9:
        return digits
    if len(digits) == 9:
        return "34" + digits
    return digits


def wa_api(path: str, body: Dict[str, Any]) -> Dict[str, Any]:
    """
    Аналог mz_wa_api: POST в Graph API.
    """
    token = getattr(settings, "WA_TOKEN", "")
    phone_id = getattr(settings, "WA_PHONE_ID", "")
    if not token or not phone_id:
        return {"ok": False, "err": "WA config missing"}

    url = f"https://graph.facebook.com/v20.0/{path}"
    try:
        r = requests.post(
            url,
            json=body,
            timeout=20,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
        )
    except Exception as e:
        log.exception("WA request error")
        return {"ok": False, "err": str(e)}

    try:
        j = r.json()
    except Exception:
        j = {"raw": r.text}

    return {
        "ok": r.status_code in (200, 201),
        "code": r.status_code,
        "resp": j,
    }


def wa_send_text(to: str, text: str) -> Dict[str, Any]:
    """
    Аналог mz_wa_send_text
    """
    to_norm = wa_msisdn(to)
    if not to_norm:
        return {"ok": False, "err": "bad_msisdn"}

    body = {
        "messaging_product": "whatsapp",
        "to": to_norm,
        "type": "text",
        "text": {
            "preview_url": True,
            "body": text,
        },
    }
    phone_id = getattr(settings, "WA_PHONE_ID", "")
    return wa_api(f"{phone_id}/messages", body)


def wa_send_document(to: str, doc_url: str, filename: str = "Documento_MEATZE.pdf") -> Dict[str, Any]:
    """
    Аналог mz_wa_send_document
    """
    to_norm = wa_msisdn(to)
    if not to_norm:
        return {"ok": False, "err": "bad_msisdn"}

    body = {
        "messaging_product": "whatsapp",
        "to": to_norm,
        "type": "document",
        "document": {
            "link": doc_url,
            "filename": filename,
        },
    }
    phone_id = getattr(settings, "WA_PHONE_ID", "")
    return wa_api(f"{phone_id}/messages", body)


def wa_send_template(
    to: str,
    tpl_name: str,
    body_params: List[str] | None = None,
    header_media: Dict[str, str] | None = None,
    lang: str | None = None,
) -> Dict[str, Any]:
    """
    Аналог mz_wa_send_template
    """
    body_params = body_params or []
    header_media = header_media or {}
    to_norm = wa_msisdn(to)
    if not to_norm:
        return {"ok": False, "err": "bad_msisdn"}

    if lang is None:
        lang = getattr(settings, "WA_LANG", "es")

    components: List[Dict[str, Any]] = []

    # header с media
    if header_media.get("type") and header_media.get("link"):
        media_type = header_media["type"]
        media_data = {"link": header_media["link"]}
        if media_type == "document" and header_media.get("filename"):
            media_data["filename"] = header_media["filename"]

        components.append(
            {
                "type": "header",
                "parameters": [
                    {
                        "type": media_type,
                        media_type: media_data,
                    }
                ],
            }
        )

    # body
    if body_params:
        components.append(
            {
                "type": "body",
                "parameters": [{"type": "text", "text": str(t)} for t in body_params],
            }
        )

    body = {
        "messaging_product": "whatsapp",
        "to": to_norm,
        "type": "template",
        "template": {
            "name": tpl_name,
            "language": {"code": lang},
            "components": components,
        },
    }
    phone_id = getattr(settings, "WA_PHONE_ID", "")
    return wa_api(f"{phone_id}/messages", body)


def wa_send_hello_world(to: str) -> Dict[str, Any]:
    return wa_send_template(to, "hello_world", [], {}, "en_US")


def wa_send_broadcast_simple(to: str, text: str) -> Dict[str, Any]:
    """
    Аналог mz_wa_send_broadcast_simple: шаблон 'meatze_broadcast_simple'
    """
    plain = text.strip()
    return wa_send_template(to, "meatze_broadcast_simple", [plain])


def wa_send_personal_txt(to: str, name: str, text: str) -> Dict[str, Any]:
    plain = text.strip()
    return wa_send_template(to, "meatze_personal_txt", [name, plain])
    
def wa_send_image(to: str, img_url: str, caption: str = "") -> Dict[str, Any]:
    """
    Отправка картинки по публичному URL (PNG/JPG).
    """
    to_norm = wa_msisdn(to)
    if not to_norm:
        return {"ok": False, "err": "bad_msisdn"}

    body = {
        "messaging_product": "whatsapp",
        "to": to_norm,
        "type": "image",
        "image": {"link": img_url},
    }
    if caption:
        body["image"]["caption"] = caption[:1024]

    phone_id = getattr(settings, "WA_PHONE_ID", "")
    return wa_api(f"{phone_id}/messages", body)


def wa_send_personal_document(to: str, name: str, text: str, doc_url: str, filename: str) -> Dict[str, Any]:
    plain = text.strip()
    return wa_send_template(
        to,
        "meatze_personal_document",
        [name, plain],
        {
            "type": "document",
            "link": doc_url,
            "filename": filename,
        },
    )


def wa_send_personal_photo(to: str, name: str, text: str, img_url: str) -> Dict[str, Any]:
    plain = text.strip()
    return wa_send_template(
        to,
        "meatze_personal_photo",
        [name, plain],
        {"type": "image", "link": img_url},
    )


def store_inbox(wa, wa_name, txt, source="meatze", direction="in"):
    wa9 = normalize_wa(wa)
    name = (wa_name or "").strip()
    WaInbox.objects.create(
        wa=wa9,
        name=name,
        source=source,
        msg=txt or "",
        direction=direction,
    )
    return wa9
    
# ========== SUBSCRIBERS / CONTACTS CRUD ==========

@api_view(["GET"])
@permission_classes([AllowAny])
@authentication_classes([])
def subscribers(request):
    """
    GET /meatze/v5/notify/subscribers
    Возвращает словарь wa -> {name, loc, active}, как в WP.
    """
    err = _require_admin(request)
    if err:
        return err

    wa_map = {}
    for c in WaContact.objects.all().order_by("wa"):
        wa_map[c.wa] = {
            "name": c.name,
            "loc": c.loc,
            "active": int(c.active),
        }

    return Response({"wa": wa_map})


@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def wa_upsert(request):
    """
    POST /meatze/v5/notify/wa-upsert
    Тело: { wa, name, loc, active }
    """
    err = _require_admin(request)
    if err:
        return err

    data = request.data or {}
    wa_raw = data.get("wa", "")
    wa = "".join(ch for ch in str(wa_raw) if ch.isdigit())
    if not wa:
        return Response({"message": "Campo 'wa' obligatorio."}, status=400)

    loc = (data.get("loc") or "").strip()
    if loc not in ("Bilbao", "Barakaldo"):
        loc = ""

    name = (data.get("name") or "").strip()
    active = int(data.get("active", 1)) != 0

    obj, _ = WaContact.objects.update_or_create(
        wa=wa,
        defaults={
            "name": name,
            "loc": loc,
            "active": active,
        },
    )

    return Response({"ok": True, "wa": obj.wa})
def wa_send_list(to: str, header: str, body: str, button_text: str, rows: list[dict], footer: str = ""):
    """
    rows: [{"id":"curso:IFCT0309", "title":"IFCT0309", "description":"Nombre curso"}...]
    """
    to_norm = wa_msisdn(to)
    if not to_norm:
        return {"ok": False, "err": "bad_msisdn"}

    payload = {
        "messaging_product": "whatsapp",
        "to": to_norm,
        "type": "interactive",
        "interactive": {
            "type": "list",
            "header": {"type": "text", "text": header[:60]},
            "body": {"text": body[:1024]},
            "action": {
                "button": button_text[:20],
                "sections": [
                    {
                        "title": "Tus cursos",
                        "rows": rows[:10],  # WhatsApp лимиты; лучше 10 за раз
                    }
                ],
            },
        },
    }
    if footer:
        payload["interactive"]["footer"] = {"text": footer[:60]}

    phone_id = getattr(settings, "WA_PHONE_ID", "")
    return wa_api(f"{phone_id}/messages", payload)

def teacher_courses(profile: UserProfile):
    qs = (Enrol.objects
          .filter(user=profile.user, role="teacher")
          .exclude(codigo="")
          .values_list("codigo", flat=True))
    codigos = sorted(set([c.strip().upper() for c in qs if c]))
    cursos = list(Curso.objects.filter(codigo__in=codigos).order_by("codigo"))
    return cursos
def build_course_rows(cursos: list[Curso]) -> list[dict]:
    rows = []
    for c in cursos:
        title = c.codigo
        desc = (c.titulo or "")[:72]
        rows.append({
            "id": f"curso:{c.codigo}",
            "title": title,
            "description": desc
        })
    return rows


@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def wa_delete(request):
    """
    POST /meatze/v5/notify/wa-delete
    { "wa": "600123123" }
    """
    err = _require_admin(request)
    if err:
        return err

    wa_raw = request.data.get("wa", "")
    wa = "".join(ch for ch in str(wa_raw) if ch.isdigit())
    if not wa:
        return Response({"message": "Campo 'wa' obligatorio."}, status=400)

    WaContact.objects.filter(wa=wa).delete()
    return Response({"ok": True})


@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def wa_toggle(request):
    """
    POST /meatze/v5/notify/wa-toggle
    — меняет active 0/1
    """
    err = _require_admin(request)
    if err:
        return err

    wa_raw = request.data.get("wa", "")
    wa = "".join(ch for ch in str(wa_raw) if ch.isdigit())
    if not wa:
        return Response({"message": "Campo 'wa' obligatorio."}, status=400)

    try:
        c = WaContact.objects.get(wa=wa)
    except WaContact.DoesNotExist:
        return Response({"message": "Contacto no encontrado."}, status=404)

    c.active = not c.active
    c.save(update_fields=["active"])
    return Response({"ok": True, "active": int(c.active)})


# ========== IMPORT (CSV) ==========

import csv
import os
from openpyxl import load_workbook
from django.core.files.uploadedfile import UploadedFile

@api_view(["POST"])
@authentication_classes([])
@permission_classes([])
def wa_import(request):
    """
    POST /meatze/v5/notify/wa-import
    form-data: file (CSV/XLSX), loc (Bilbao|Barakaldo|'')
    """
    err = _require_admin(request)
    if err:
        return err

    up = request.FILES.get("file")
    if not up:
        return Response({"ok": False, "message": "Falta fichero"}, status=400)

    loc = (request.POST.get("loc") or "").strip()
    if loc not in ("Bilbao", "Barakaldo"):
        loc = ""

    filename = (up.name or "")
    ext = os.path.splitext(filename)[1].lower()

    inserted = 0
    updated = 0
    skipped = 0   # ← теперь точно объявлен ДО всех веток

    rows = []

    # ---------- XLSX ----------
    if ext == ".xlsx":
        wb = load_workbook(up, read_only=True, data_only=True)
        ws = wb.active

        # Строка 1 – служебная (код курса + название), начинаем с 2
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # ожидаем: A = índice, B = nombre, C = teléfono
            if len(row) < 3:
                # служебная/пустая – пропускаем
                continue

            idx_raw, name_raw, phone_raw = row[0], row[1], row[2]

            # Если нет телефона – это либо пустая строка, либо служебная → пропускаем
            if phone_raw in (None, ""):
                continue

            wa = normalize_wa(phone_raw)
            if not wa or len(wa) != 9:
                # не похоже на нормальный номер → считаем служебной/битой
                skipped += 1
                continue

            name_str = (str(name_raw or "").strip() or "Sin nombre")

            rows.append({
                "wa": wa,
                "name": name_str,
                "loc": loc,
                "active": 1,
            })

    # ---------- CSV / TXT ----------
    elif ext in (".csv", ".txt"):
        try:
            content = up.read()
            try:
                text = content.decode("utf-8")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
        except Exception:
            return Response(
                {"ok": False, "message": "No se puede leer el fichero"},
                status=400,
            )

        buf = io.StringIO(text, newline="")

        # пытаемся угадать разделитель, но по умолчанию ставим ';'
        try:
            sample = text[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"

        reader = csv.reader(buf, dialect)

        for row in reader:
            if not row:
                continue

            vals = [(c or "").strip() for c in row]
            # служебные строки без цифр/телефонов – пропускаем
            if not any(ch.isdigit() for ch in "".join(vals)):
                continue

            if len(vals) == 1:
                skipped += 1
                continue
            elif len(vals) == 2:
                name_raw, phone_raw = vals
            else:
                # берём две последние как name/phone (как в WP)
                name_raw, phone_raw = vals[-2], vals[-1]

            wa = normalize_wa(phone_raw)
            if not wa or len(wa) != 9:
                skipped += 1
                continue

            name_str = (name_raw or "").strip() or "Sin nombre"

            rows.append({
                "wa": wa,
                "name": name_str,
                "loc": loc,
                "active": 1,
            })

    else:
        return Response(
            {"ok": False, "message": "Formato no soportado (usa CSV o XLSX)"},
            status=400,
        )

    # ---------- сохраняем в БД ----------
    for r in rows:
        wa = r["wa"]
        name_str = r["name"]
        active = bool(r.get("active", 1))

        obj, created = WaContact.objects.update_or_create(
            wa=wa,
            defaults={
                "name": name_str,
                "loc": loc,
                "active": active,
            },
        )
        if created:
            inserted += 1
        else:
            updated += 1

    return Response(
        {
            "ok": True,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
        }
    )



# ========== UPLOAD MEDIA (для WA) ==========

@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def upload_wa(request):
    """
    POST /meatze/v5/notify/upload-wa
    form-data: file
    Возвращает url + filename, как WP upload-wa.
    """
    err = _require_admin(request)
    if err:
        return err

    f = request.FILES.get("file")
    if not f:
        return Response({"ok": False, "message": "Fichero requerido."}, status=400)

    filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{f.name}"
    path = default_storage.save(f"wa_broadcast/{filename}", ContentFile(f.read()))
    file_url = request.build_absolute_uri(default_storage.url(path))

    return Response({"ok": True, "url": file_url, "filename": f.name})


# ========== BROADCAST (главная кнопка рассылки) ==========

@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def broadcast(request):
    """
    POST /meatze/v5/notify/broadcast
    Тело максимально похоже на WP /news/broadcast, но реализуем только канал 'wa'.

    Важные поля:
      - mode: "all" | "selected"
      - channels: ["wa"]
      - text: текст рассылки
      - sel_wa: ["600...", "699..."] (для mode=selected)
      - test_wa: "600..." — если есть, шлём только ему
      - wa_tpl: "personal_txt" | "personal_photo" | "personal_doc" | "hello_world" | "broadcast"
      - wa_media_url, wa_media_name
      - wa_loc: "Bilbao" | "Barakaldo" | ""
    """
    err = _require_admin(request)
    if err:
        return err

    b = request.data or {}
    mode = b.get("mode") or "all"
    channels = list(b.get("channels") or [])
    text = (b.get("text") or "").strip()
    sel_wa = list(b.get("sel_wa") or [])
    test_wa = "".join(ch for ch in str(b.get("test_wa") or "") if ch.isdigit())
    wa_tpl = b.get("wa_tpl") or "broadcast"
    wa_media_url = (b.get("wa_media_url") or "").strip()
    wa_media_name = (b.get("wa_media_name") or "").strip() or "Ficha_MEATZE.pdf"
    wa_loc = (b.get("wa_loc") or "").strip()
    if wa_loc not in ("Bilbao", "Barakaldo"):
        wa_loc = ""

    if not text:
        return Response({"ok": False, "message": "Texto obligatorio."}, status=400)

    # По факту нас интересует только WA канал
    if channels and "wa" not in channels:
        return Response({"ok": False, "message": "Sólo canal 'wa' soportado de momento."}, status=400)

    # Собираем список получателей
    if test_wa:
        wa_list = [
            {"wa": test_wa, "name": "", "loc": "", "active": True},
        ]
    else:
        qs = WaContact.objects.filter(active=True)
        if wa_loc:
            qs = qs.filter(loc=wa_loc)
        wa_list = list(qs.values("wa", "name", "loc", "active"))

        if mode == "selected" and sel_wa:
            sel_set = {str("".join(ch for ch in w if ch.isdigit())) for w in sel_wa}
            wa_list = [w for w in wa_list if w["wa"] in sel_set]

    wa_ok = 0
    wa_fail = 0
    wa_last = None

    for w in wa_list:
        num = w.get("wa")
        num_digits = "".join(ch for ch in str(num) if ch.isdigit())
        if not num_digits:
            wa_fail += 1
            continue

        name = (w.get("name") or "").strip() or "alumno/a"

        if wa_tpl == "personal_txt":
            res = wa_send_personal_txt(num_digits, name, text)

        elif wa_tpl == "personal_photo":
            if not wa_media_url:
                wa_fail += 1
                continue
            res = wa_send_personal_photo(num_digits, name, text, wa_media_url)

        elif wa_tpl == "personal_doc":
            if not wa_media_url:
                wa_fail += 1
                continue
            # полностью повторять WP-логику с двухэтапной отправкой можно,
            # но проще сразу использовать шаблон с документом:
            res = wa_send_personal_document(num_digits, name, text, wa_media_url, wa_media_name)

        elif wa_tpl == "hello_world":
            res = wa_send_hello_world(num_digits)

        else:  # broadcast simple
            res = wa_send_broadcast_simple(num_digits, text)

        wa_last = res
        if res.get("ok"):
            wa_ok += 1
        else:
            wa_fail += 1

        if test_wa:
            break

    return Response(
        {
            "ok": True,
            "stats": {
                "wa_ok": wa_ok,
                "wa_fail": wa_fail,
                "wa_last": wa_last,
            },
        }
    )

def _send_teacher_course_pack(request, to_wa: str, codigo: str):
    codigo = (codigo or "").strip().upper()
    cobj = Curso.objects.filter(codigo=codigo).first()
    titulo = (cobj.titulo or "").strip() if cobj else ""
    header = f"*{codigo}*" + (f" — {titulo}" if titulo else "")

    # span из Horario
    start, end = _course_span_from_horario(codigo)
    start_txt = start.isoformat() if start else "—"
    end_txt   = end.isoformat() if end else "—"

    # tramos
    segs = _course_time_segments(codigo)

    # месяцы только по расписанию
    months = _course_months_by_horario(codigo)
    if not months:
        wa_send_text(to_wa, header + "\n— No hay horario para este curso.")
        return

    lines = [
      header,
      f"Fechas: {start_txt} → {end_txt}",
      "",
      "📅 Calendario por meses (como export):",
      "Leyenda: en cada mes, la hora indica HH:MM–HH:MM y el texto entre [ ] son los primeros 10 caracteres del módulo.",
      "",
    ]
    for (yy, mm) in months:
        lines += _month_summary_lines(codigo, yy, mm)
    wa_send_text(to_wa, "\n".join(lines)[:3800])

    # PNG на все месяцы
    for (yy, mm) in months:
        _send_teacher_month_calendar(request, to_wa, codigo, yy, mm)


# ========== WEBHOOK (входящие + рассылка админу по разным номерам) ==========

@api_view(["GET", "POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def ws_webhook(request):
    # === Verify (GET) ===
    if request.method == "GET":
        mode = request.GET.get("hub.mode")
        token = request.GET.get("hub.verify_token")
        chal = request.GET.get("hub.challenge")

        if mode == "subscribe" and token == getattr(settings, "WA_VERIFY_TOKEN", "") and chal:
            # ВАЖНО: вернуть голый текст, без JSON
            return HttpResponse(chal, content_type="text/plain", status=200)

        return Response({"error": "verify mismatch"}, status=403)

    # === Incoming (POST) ===
    payload = request.data or {}
    try:
        chg = payload["entry"][0]["changes"][0]["value"]
        msg = chg.get("messages", [None])[0]
    except Exception:
        return Response({"ok": True})

    if not msg:
        return Response({"ok": True})

    from_ = msg.get("from") or ""
    if not from_:
        return Response({"ok": True})

    profile = (chg.get("contacts", [{}])[0].get("profile") or {}).get("name", "") or ""

    msg_type = msg.get("type") or ""
    txt = ""
    selected_id = ""

    # --- interactive ---
    if msg_type == "interactive":
        inter = msg.get("interactive") or {}
        lr = inter.get("list_reply") or {}
        br = inter.get("button_reply") or {}
        if lr:
            selected_id = (lr.get("id") or "").strip()
            txt = (lr.get("title") or "").strip()
        elif br:
            selected_id = (br.get("id") or "").strip()
            txt = (br.get("title") or "").strip()

    # --- text/media ---
    if msg_type == "text":
        txt = msg.get("text", {}).get("body", "") or ""
    elif msg_type == "image":
        caption = msg.get("image", {}).get("caption", "") or ""
        txt = "🖼 Imagen recibida" + (": " + caption if caption else "")
    elif msg_type == "document":
        filename = msg.get("document", {}).get("filename", "") or ""
        caption = msg.get("document", {}).get("caption", "") or ""
        parts = ["📎 Documento"]
        if filename:
            parts.append(filename)
        if caption:
            parts.append("— " + caption)
        txt = " ".join(parts)
    elif msg_type == "interactive":
        # txt уже установлен выше
        pass
    else:
        txt = f"[mensaje de tipo {msg_type or 'desconocido'}]"
        
        
    # ===== TEACHER FAST LANE (НЕ пишем в WaInbox и НЕ пересылаем админу) =====
    tprof = _teacher_profile_by_from(from_)
    if tprof:
        cursos = teacher_courses(tprof)
        codes = [c.codigo for c in cursos]

        # 1) если нажали курс — обрабатываем ВСЕГДА
        if selected_id.startswith("curso:"):
            codigo = selected_id.split(":", 1)[1].strip().upper()
            if codigo in codes:
                wa_send_text(from_, _teacher_course_info_text(request, codigo))
                _send_teacher_course_pack(request, from_, codigo)
            return Response({"ok": True})

        # 2) если прислали код текстом — обрабатываем ВСЕГДА
        t = (txt or "").strip().upper()
        if t and t in codes:
            wa_send_text(from_, _teacher_course_info_text(request, t))
            _send_teacher_course_pack(request, from_, t)
            return Response({"ok": True})

        # 3) иначе — приветствие + список курсов (кешируем, чтобы не спамить)
        wa9 = normalize_wa(from_)
        key = f"wa:t:{wa9}:hello"
        if not cache.get(key):
            cache.set(key, 1, 30)  # 30с, чтобы не повторять каждое сообщение

            link = request.build_absolute_uri("/acceder/?tab=profile")
            name = (tprof.display_name or tprof.build_display_name() or "profe").strip()
            header = "MEATZE · Docentes"
            body = f"Hola {name} 👋\nPerfil: {link}\n\nTus cursos (elige uno):"
            rows = build_course_rows(cursos)

            if rows:
                wa_send_list(from_, header, body, "Cursos", rows, footer="Te envío meses + PNG")
            else:
                wa_send_text(from_, _teacher_welcome_text(request, tprof, []))

        return Response({"ok": True})


    # ===== NO-TEACHER FLOW =====
    wa9 = store_inbox(from_, profile, txt, source="meatze", direction="in")
    # определяем loc по WaContact
    from_plain = "".join(ch for ch in str(from_) if ch.isdigit())
    wa_short = from_plain[2:] if from_plain.startswith("34") and len(from_plain) > 9 else from_plain

    loc = (
        WaContact.objects.filter(wa=wa_short).values_list("loc", flat=True).first()
        or ""
    )

    admin_wa = ""
    if loc == "Barakaldo":
        admin_wa = getattr(settings, "WA_ADMIN_BARAKALDO", "")
    elif loc == "Bilbao":
        admin_wa = getattr(settings, "WA_ADMIN_BILBAO", "")

    if admin_wa:
        prev = txt[:400] or "[mensaje vacío]"
        admin_msg = (
            "Nuevo mensaje recibido en *MEATZE*:\n\n"
            f"De: {profile or 'Sin nombre'}\n"
            f"Número: +{from_plain}\n"
        )
        if loc:
            admin_msg += f"Localidad: {loc}\n"
        admin_msg += "\n" + prev
        wa_send_text(admin_wa, admin_msg)

    return Response({"ok": True})

@api_view(["POST"])
@authentication_classes([])     # токен мы читаем сами
@permission_classes([])         # без DRF-permissions, только наш guard
def wa_clear(request):
    """
    POST /meatze/v5/notify/wa-clear
    Полностью очищает таблицу WhatsApp-контактов.
    """
    err = _require_admin(request)
    if err:
        return err

    with transaction.atomic():
        deleted, _ = WaContact.objects.all().delete()

    return Response({"ok": True, "deleted": deleted})

@api_view(["GET"])
@authentication_classes([])
@permission_classes([AllowAny])
def wa_inbox(request):
    """
    GET /meatze/v5/notify/wa-inbox?limit=50
    Отдаём последние N сообщений для панели чата.
    Формат под wa.js: { items: [ {id, wa, msg, created_at, wa_name, sub_name, loc, direction}, ... ] }
    """
    err = _require_admin(request)
    if err:
        return err

    try:
        limit = int(request.GET.get("limit") or 50)
    except ValueError:
        limit = 50
    if limit > 200:
        limit = 200

    qs = WaInbox.objects.all().order_by("-created_at")[:limit]

    # подцепим имена/loc из WaContact
    wa_list = list({r.wa for r in qs})
    contacts = WaContact.objects.filter(wa__in=wa_list)
    by_wa = {c.wa: c for c in contacts}

    items = []
    for r in qs:
        c = by_wa.get(r.wa)
        items.append({
            "id": r.id,
            "wa": r.wa,
            "msg": r.msg,
            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "wa_name": r.name,              # имя из WA профиля (если есть)
            "sub_name": c.name if c else "",# имя из нашей базы контактов
            "loc": c.loc if c else "",      # Bilbao / Barakaldo
            "direction": r.direction,       # 'in' / 'out'
        })

    return Response({"items": items})

@api_view(["POST"])
@authentication_classes([])
@permission_classes([AllowAny])
def wa_inbox_delete(request):
    """
    POST /meatze/v5/notify/wa-inbox-delete
    { "wa": "600123123" }
    """
    err = _require_admin(request)
    if err:
        return err

    wa_raw = request.data.get("wa", "")
    wa = re.sub(r"\D+", "", str(wa_raw or ""))
    if not wa:
        return Response({"message": "Campo 'wa' obligatorio."}, status=400)

    WaInbox.objects.filter(wa=wa).delete()
    return Response({"ok": True})

@api_view(["POST"])
@authentication_classes([])
@permission_classes([AllowAny])
def wa_reply(request):
    """
    POST /meatze/v5/notify/wa-reply
    { "wa": "600123123", "text": "..." }
    Отправляет текст в WhatsApp и пишет запись в WaInbox (direction='out').
    """
    err = _require_admin(request)
    if err:
        return err

    data = request.data or {}
    wa_raw = data.get("wa", "")
    text = (data.get("text") or "").strip()

    if not wa_raw or not text:
        return Response(
            {"message": "Campos 'wa' y 'text' obligatorios."},
            status=400,
        )

    wa_short = normalize_wa(wa_raw)         # 9 цифр
    if not wa_short:
        return Response({"message": "Número inválido."}, status=400)

    # Отправляем через уже существующий helper
    res = wa_send_text(wa_short, text)

    # Логируем исходящее в inbox
    store_inbox(wa_short, "", text, source="meatze", direction="out")

    if not res.get("ok"):
        return Response(
            {"ok": False, "error": res.get("err") or "WA error", "resp": res},
            status=502,
        )

    return Response({"ok": True})

# пример — положи рядом с другими v5 AI-ручками
import requests
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.conf import settings


OLLAMA_URL = getattr(settings, "OLLAMA_URL", "http://127.0.0.1:11434/api/chat")
OLLAMA_MODEL = getattr(settings, "OLLAMA_MODEL", "llama3.1")


@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def ai_portal_helper(request):
    """
    POST /meatze/v5/ai/portal_helper
    { "question": "...", "history": [ { "role": "user"|"assistant", "content": "..." }, ... ] }
    """
    question = (request.data.get("question") or "").strip()
    history = request.data.get("history") or []

    if not question:
        return Response({"message": "Pregunta vacía."}, status=400)

    # safety: берём только последние 8 сообщений и нормализуем структуру
    msgs = []
    for m in history[-8:]:
        role = m.get("role")
        content = (m.get("content") or "").strip()
        if role in ("user", "assistant") and content:
            msgs.append({"role": role, "content": content})

    # --- ДОБАВЛЯЕМ RAG-КОНТЕКСТ ---
    kb_context = ""
    try:
        # top_k можешь менять (3–8 обычно норм)
        kb_context = retrieve_context(question, top_k=6)
    except Exception as e:
        # если что-то сломалось, просто работаем без базы, но не падаем
        kb_context = ""

    system_prompt = (
        "Eres *Asistente MEATZE Campus*, un ayudante virtual del portal de formación MEATZE.\n"
        "- Respondes SIEMPRE en español (puedes aclarar algo en euskera o ruso si el usuario lo usa).\n"
        "- SOLO puedes hablar de MEATZE, del portal alumno, de los cursos subvencionados, módulos, "
        "calendario, materiales, chat, IA, acceso, contraseñas, etc.\n"
        "- NO puedes inventar botones, secciones, apps móviles ni formularios de registro que NO estén "
        "descritos en el CONTEXTO que recibes.\n"
        "- Si en el contexto no aparece una funcionalidad (por ejemplo 'Registro', 'app móvil'), debes "
        "decir claramente que el portal NO tiene eso y explicar cómo se accede realmente (por PIN, "
        "correo del centro, etc.).\n"
        "- Si te preguntan algo que no sea sobre MEATZE o sobre este portal, respondes brevemente "
        "que este asistente solo puede ayudar con el campus MEATZE y propones alguna acción dentro del portal.\n"
        "- Sé claro, breve y práctico. Ejemplos: cómo entrar, cómo ver calendario, cómo funcionan los módulos, "
        "qué es la pestaña 'IA', cómo contactar por WhatsApp, etc.\n"
    )

    # Вшиваем контекст базы в промпт
    if kb_context:
        system_prompt += (
            "\n\n[CONTEXTO TÉCNICO DEL PORTAL MEATZE]\n"
            "Usa SOLO la siguiente información para responder. Si algo no aparece aquí, responde que "
            "no lo sabes o que no existe en el portal y recomienda contactar con el centro MEATZE.\n\n"
            f"{kb_context}\n"
        )

    payload = {
        "model": OLLAMA_MODEL,
        "messages": (
            [{"role": "system", "content": system_prompt}]
            + msgs
            + [{"role": "user", "content": question}]
        ),
        "stream": False,
    }

    try:
        r = requests.post(OLLAMA_URL, json=payload, timeout=60)
        r.raise_for_status()
        data = r.json()

        content = ""
        if isinstance(data, dict):
            if "message" in data:
                content = data["message"].get("content", "")
            elif "choices" in data:
                content = data["choices"][0]["message"]["content"]

        content = (content or "").strip()
        if not content:
            content = "Lo siento, ahora mismo no puedo responder. Intenta de nuevo en un momento."

        return Response({"answer": content})
    except requests.RequestException as e:
        return Response(
            {"message": "Error al contactar con el motor IA.", "detail": str(e)},
            status=502,
        )
